import openpyxl

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Activate the first sheet
sheet = workbook.active

# Rename the first sheet
sheet.title = "sheet_1"

# Save the workbook
workbook.save("excel_sheet.xlsx")

# use command to open ->  open excel_sheet.xlsx

from openpyxl import load_workbook, Workbook

wb = load_workbook('excel_sheet.xlsx')
new_ws = wb.create_sheet('Sheet2')
new_ws['A1'] = 'Hello, world A1!'
new_ws = wb.create_sheet('Sheet3')
new_ws['B1'] = 'Hello, world B1!'
wb.save('excel_sheet.xlsx')
